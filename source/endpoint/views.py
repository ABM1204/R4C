from django.http import JsonResponse
from django.shortcuts import render
import logging
import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Sum
from robots.models import Robot, RobotProduction
from django.utils import timezone

def index(request):
    return render(request, 'index.html')

def get_all_robots(request):
    if request.method == "GET":
        try:
            robots = Robot.objects.all()
            robots_data = [
                {
                    "id": robot.id,
                    "serial": robot.serial,
                    "model": robot.model,
                    "version": robot.version,
                    "created": robot.created.strftime("%Y-%m-%d %H:%M:%S")
                }
                for robot in robots
            ]
            return JsonResponse({"status": "success", "data": robots_data}, status=200)
        except Exception as e:
            return JsonResponse({"status": "error", "message": "An unexpected error occurred."}, status=500)
    return JsonResponse({"status": "error", "message": "Only GET requests are allowed."}, status=405)


logger =  logging.getLogger(__name__)

def generate_robot_summary(request):
    try:
        today = timezone.now().date()
        start_of_week = today - datetime.timedelta(days=today.weekday())
        start_of_week = timezone.make_aware(datetime.datetime.combine(start_of_week, datetime.time.min))

        logger.debug(f"Start of week: {start_of_week}")

        wb = create_robot_summary_excel(start_of_week)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="robot_production_summary.xlsx"'

        wb.save(response)

        logger.debug("Excel file successfully sent in response.")
        return response

    except Exception as e:
        logger.exception(f"Error generating robot summary: {e}")
        return JsonResponse({"status": "error", "message": f"Error generating robot summary: {e}"}, status=500)


def create_robot_summary_excel(start_of_week):
    wb = Workbook()
    ws = wb.active

    models = Robot.objects.all()
    logger.debug(f"Found {len(models)} robot models in the database.")

    if not models.exists():
        logger.warning("No robot models found in the database.")
        ws.append(["Нет данных о моделях роботов"])
        return wb

    for model in models:
        ws = wb.create_sheet(title=model.model[:31])

        ws.append(["Модель", "Версия", "Количество за неделю"])

        robot_versions = RobotProduction.objects.filter(
            robot__model=model,
            produced_at__gte=start_of_week
        ).values('robot__version').annotate(
            total_produced=Sum('quantity')
        ).order_by('robot__version')

        if not robot_versions.exists():
            logger.info(f"No production data found for model {model.model}.")
            ws.append([model.model, "Нет данных", 0])
            continue

        for version in robot_versions:
            ws.append([model.model, version['robot__version'], version['total_produced']])

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb